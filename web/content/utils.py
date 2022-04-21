from openpyxl import load_workbook
from content.models import Trivia, NxSong, Movie, Person
from django.db.models import Q

def iter_rows(ws):
     for row in ws.iter_rows():
         yield [cell.value for cell in row]


def reload_trivia(file='./utils/FINAL-updated active list-17 sept-18.xlsx'):
    wb = load_workbook(file, read_only=True)
    ws = wb['Sheet1']
    rows = iter_rows(ws)

    for r in rows:
        if r[0] and r[1] and type(r[2])==int:
            n = NxSong.objects.filter(content_id=r[0]).first()
            if n:
                n.content_id=r[1]
                m = r[4]
                if m and type(m)==str and m.strip():
                    movie = Movie.objects.filter(Q(year_of_release='{}-01-01'.format(r[2]))&Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        n.movie=movie

                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                n.actors.add(pr)
                n.save()
            else:
                nx = NxSong.objects.create(content_id=r[0], title=r[3].strip().title(), released_on='{}-01-01'.format(r[2]))
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        nx.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                nx.actors.add(pr)
                nx.save()
            nx = NxSong.objects.filter(content_id=r[1]).first()
            if nx:
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        nx.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                nx.actors.add(pr)
                nx.save()
            else:
                nx = NxSong.objects.create(content_id=r[1], title=r[3].strip().title(), released_on='{}-01-01'.format(r[2]))
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        nx.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                nx.actors.add(pr)
                nx.save()
        elif r[1] and type(r[2])==int:
            nx = NxSong.objects.filter(content_id=r[1]).first()
            if nx:
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        nx.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                nx.actors.add(pr)
                nx.save()
            else:
                nx = NxSong.objects.create(content_id=r[1], title=r[3].strip().title(), released_on='{}-01-01'.format(r[2]))
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        nx.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                nx.actors.add(pr)
                nx.save()
        elif r[0] and type(r[2])==int:
            n = NxSong.objects.filter(content_id=r[0]).first()
            if n:
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        n.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                n.actors.add(pr)
                n.save()
            else:
                nx = NxSong.objects.create(content_id=r[0], title=r[3].strip().title(), released_on='{}-01-01'.format(r[2]))
                m = r[4]
                if m and type(m) == str and m.strip():
                    movie = Movie.objects.filter(
                        Q(year_of_release='{}-01-01'.format(r[2])) & Q(movie_title=r[4].strip().title())).first()
                    if movie:
                        nx.movie = movie
                persons = r[7:len(r)]
                for p in persons:
                    if p:
                        person = Person.objects.filter(name=p.strip().title())
                        if person:
                            for pr in person:
                                nx.actors.add(pr)
                nx.save()